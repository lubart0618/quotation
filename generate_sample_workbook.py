from __future__ import annotations

from copy import copy
from datetime import date, timedelta

from openpyxl import Workbook


def add_headers(ws, headers):
    ws.append(headers)
    for cell in ws[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font


def main() -> None:
    wb = Workbook()

    ws_employees = wb.active
    ws_employees.title = "Employees"
    add_headers(
        ws_employees,
        [
            "employee_id",
            "name",
            "groups",
            "available_days",
            "available_slots",
            "weekly_max_shifts",
            "priority",
            "notes",
        ],
    )

    employees = [
        ("E001", "Amy", "electronics,liquor", "Mon,Tue,Wed,Thu,Fri,Sat", "AM,PM", 5, 3, ""),
        ("E002", "Ben", "snacks,candy", "Tue,Wed,Thu,Fri,Sat,Sun", "AM,PM", 5, 2, ""),
        ("E003", "Cara", "liquor", "Fri,Sat,Sun", "PM", 3, 4, ""),
        ("E004", "David", "electronics", "Mon,Tue,Wed,Thu,Fri", "AM", 5, 2, ""),
        ("E005", "Ella", "snacks,candy,electronics", "Mon,Tue,Wed,Thu,Fri,Sat,Sun", "AM,PM", 6, 5, ""),
        ("E006", "Frank", "liquor,snacks", "Thu,Fri,Sat,Sun", "AM,PM", 4, 1, ""),
    ]
    for row in employees:
        ws_employees.append(row)

    ws_leaves = wb.create_sheet("Leaves")
    add_headers(ws_leaves, ["employee_id", "date", "slot", "reason"])
    ws_leaves.append(["E005", "2026-04-15", "PM", "personal leave"])
    ws_leaves.append(["E001", "2026-04-16", "AM", "training"])

    ws_events = wb.create_sheet("Events")
    add_headers(
        ws_events,
        [
            "event_id",
            "date",
            "slot",
            "store",
            "brand",
            "category",
            "required_staff",
            "preferred_group",
            "notes",
        ],
    )

    start = date(2026, 4, 14)
    rows = []
    categories = [
        ("electronics", "Sony"),
        ("liquor", "Asahi"),
        ("snacks", "Lay's"),
        ("candy", "KitKat"),
    ]
    for index in range(10):
        event_date = start + timedelta(days=index % 5)
        slot = "AM" if index % 2 == 0 else "PM"
        category, brand = categories[index % len(categories)]
        rows.append(
            [
                f"EV{index + 1:03d}",
                event_date.isoformat(),
                slot,
                f"Store-{(index % 3) + 1}",
                brand,
                category,
                1 if index < 8 else 2,
                category,
                "",
            ]
        )

    for row in rows:
        ws_events.append(row)

    wb.save("input_template.xlsx")
    print("Created input_template.xlsx")


if __name__ == "__main__":
    main()
