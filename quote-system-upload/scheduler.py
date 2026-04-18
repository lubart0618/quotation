from __future__ import annotations

import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


VALID_EVENT_SLOTS = {"AM", "PM"}
VALID_LEAVE_SLOTS = {"AM", "PM", "FULL"}
WEEKDAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


@dataclass
class Employee:
    employee_id: str
    name: str
    groups: set[str]
    available_days: set[str]
    available_slots: set[str]
    weekly_max_shifts: int
    priority: int
    notes: str = ""
    assigned_events: list[str] = field(default_factory=list)


@dataclass
class LeaveRecord:
    employee_id: str
    date: str
    slot: str
    reason: str = ""


@dataclass
class Event:
    event_id: str
    date: str
    slot: str
    store: str
    brand: str
    category: str
    required_staff: int
    preferred_group: str
    notes: str = ""


def normalize_csv_set(value: object) -> set[str]:
    if value is None:
        return set()
    parts = [str(part).strip() for part in str(value).split(",")]
    return {part for part in parts if part}


def required_headers(ws, headers: Iterable[str]) -> dict[str, int]:
    found = {str(cell.value).strip(): index for index, cell in enumerate(ws[1], start=1) if cell.value}
    missing = [header for header in headers if header not in found]
    if missing:
        raise ValueError(f"Sheet '{ws.title}' is missing required columns: {', '.join(missing)}")
    return found


def parse_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    text = str(value).strip()
    datetime.strptime(text, "%Y-%m-%d")
    return text


def weekday_from_date(date_text: str) -> str:
    return WEEKDAY_NAMES[datetime.strptime(date_text, "%Y-%m-%d").weekday()]


def load_input_workbook(path: Path) -> tuple[dict[str, Employee], list[LeaveRecord], list[Event]]:
    wb = load_workbook(path)

    employees_ws = wb["Employees"]
    leaves_ws = wb["Leaves"]
    events_ws = wb["Events"]

    employee_headers = required_headers(
        employees_ws,
        [
            "employee_id",
            "name",
            "groups",
            "available_days",
            "available_slots",
            "weekly_max_shifts",
            "priority",
            "notes",
        ],
    )

    leave_headers = required_headers(leaves_ws, ["employee_id", "date", "slot", "reason"])
    event_headers = required_headers(
        events_ws,
        [
            "event_id",
            "date",
            "slot",
            "store",
            "brand",
            "category",
            "required_staff",
            "preferred_group",
            "notes",
        ],
    )

    employees: dict[str, Employee] = {}
    for row in employees_ws.iter_rows(min_row=2, values_only=False):
        employee_id = str(row[employee_headers["employee_id"] - 1].value or "").strip()
        if not employee_id:
            continue
        employees[employee_id] = Employee(
            employee_id=employee_id,
            name=str(row[employee_headers["name"] - 1].value or "").strip(),
            groups={group.lower() for group in normalize_csv_set(row[employee_headers["groups"] - 1].value)},
            available_days=normalize_csv_set(row[employee_headers["available_days"] - 1].value),
            available_slots={slot.upper() for slot in normalize_csv_set(row[employee_headers["available_slots"] - 1].value)},
            weekly_max_shifts=int(row[employee_headers["weekly_max_shifts"] - 1].value or 0),
            priority=int(row[employee_headers["priority"] - 1].value or 0),
            notes=str(row[employee_headers["notes"] - 1].value or "").strip(),
        )

    leaves: list[LeaveRecord] = []
    for row in leaves_ws.iter_rows(min_row=2, values_only=False):
        employee_id = str(row[leave_headers["employee_id"] - 1].value or "").strip()
        if not employee_id:
            continue
        slot = str(row[leave_headers["slot"] - 1].value or "").strip().upper()
        if slot not in VALID_LEAVE_SLOTS:
            raise ValueError(f"Invalid leave slot '{slot}' for employee {employee_id}")
        leaves.append(
            LeaveRecord(
                employee_id=employee_id,
                date=parse_date(row[leave_headers["date"] - 1].value),
                slot=slot,
                reason=str(row[leave_headers["reason"] - 1].value or "").strip(),
            )
        )

    events: list[Event] = []
    for row in events_ws.iter_rows(min_row=2, values_only=False):
        event_id = str(row[event_headers["event_id"] - 1].value or "").strip()
        if not event_id:
            continue
        slot = str(row[event_headers["slot"] - 1].value or "").strip().upper()
        if slot not in VALID_EVENT_SLOTS:
            raise ValueError(f"Invalid event slot '{slot}' for event {event_id}")
        category = str(row[event_headers["category"] - 1].value or "").strip().lower()
        preferred_group = str(row[event_headers["preferred_group"] - 1].value or "").strip().lower() or category
        events.append(
            Event(
                event_id=event_id,
                date=parse_date(row[event_headers["date"] - 1].value),
                slot=slot,
                store=str(row[event_headers["store"] - 1].value or "").strip(),
                brand=str(row[event_headers["brand"] - 1].value or "").strip(),
                category=category,
                required_staff=int(row[event_headers["required_staff"] - 1].value or 0),
                preferred_group=preferred_group,
                notes=str(row[event_headers["notes"] - 1].value or "").strip(),
            )
        )

    return employees, leaves, events


def build_leave_lookup(leaves: list[LeaveRecord]) -> dict[tuple[str, str], set[str]]:
    lookup: dict[tuple[str, str], set[str]] = {}
    for leave in leaves:
        key = (leave.employee_id, leave.date)
        lookup.setdefault(key, set()).add(leave.slot)
    return lookup


def is_on_leave(leave_lookup: dict[tuple[str, str], set[str]], employee_id: str, date: str, slot: str) -> bool:
    slots = leave_lookup.get((employee_id, date), set())
    return "FULL" in slots or slot in slots


def is_available(employee: Employee, event: Event, leave_lookup: dict[tuple[str, str], set[str]], occupied: set[tuple[str, str, str]]) -> bool:
    weekday = weekday_from_date(event.date)
    if event.preferred_group not in employee.groups and event.category not in employee.groups:
        return False
    if weekday not in employee.available_days:
        return False
    if event.slot not in employee.available_slots and "FULL" not in employee.available_slots:
        return False
    if is_on_leave(leave_lookup, employee.employee_id, event.date, event.slot):
        return False
    if (employee.employee_id, event.date, event.slot) in occupied:
        return False
    if len(employee.assigned_events) >= employee.weekly_max_shifts:
        return False
    return True


def candidate_score(employee: Employee, event: Event) -> tuple[int, int, int, str]:
    matching_groups = int(event.preferred_group in employee.groups) + int(event.category in employee.groups)
    return (
        len(employee.assigned_events),
        -employee.priority,
        -matching_groups,
        employee.employee_id,
    )


def schedule(employees: dict[str, Employee], leaves: list[LeaveRecord], events: list[Event]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    leave_lookup = build_leave_lookup(leaves)
    occupied: set[tuple[str, str, str]] = set()

    assignments: list[dict[str, str]] = []
    unfilled: list[dict[str, str]] = []

    sorted_events = sorted(events, key=lambda e: (e.date, e.slot, e.store, e.brand, e.event_id))
    for event in sorted_events:
        for seat in range(1, event.required_staff + 1):
            eligible = [
                employee
                for employee in employees.values()
                if is_available(employee, event, leave_lookup, occupied)
            ]

            if not eligible:
                unfilled.append(
                    {
                        "event_id": event.event_id,
                        "date": event.date,
                        "slot": event.slot,
                        "store": event.store,
                        "brand": event.brand,
                        "category": event.category,
                        "missing_seat": str(seat),
                        "reason": "No eligible employee found",
                    }
                )
                continue

            chosen = min(eligible, key=lambda employee: candidate_score(employee, event))
            chosen.assigned_events.append(event.event_id)
            occupied.add((chosen.employee_id, event.date, event.slot))
            assignments.append(
                {
                    "event_id": event.event_id,
                    "date": event.date,
                    "slot": event.slot,
                    "store": event.store,
                    "brand": event.brand,
                    "category": event.category,
                    "employee_id": chosen.employee_id,
                    "employee_name": chosen.name,
                    "seat_no": str(seat),
                }
            )

    return assignments, unfilled


def build_summary(employees: dict[str, Employee]) -> list[dict[str, str]]:
    rows = []
    for employee in sorted(employees.values(), key=lambda item: item.employee_id):
        rows.append(
            {
                "employee_id": employee.employee_id,
                "name": employee.name,
                "assigned_shifts": str(len(employee.assigned_events)),
                "weekly_max_shifts": str(employee.weekly_max_shifts),
                "groups": ", ".join(sorted(employee.groups)),
            }
        )
    return rows


def run_scheduler(input_path: Path, output_path: Path) -> tuple[list[dict[str, str]], list[dict[str, str]], list[dict[str, str]]]:
    employees, leaves, events = load_input_workbook(input_path)
    assignments, unfilled = schedule(employees, leaves, events)
    summary = build_summary(employees)
    write_output(output_path, assignments, unfilled, summary)
    return assignments, unfilled, summary


def write_output(path: Path, assignments: list[dict[str, str]], unfilled: list[dict[str, str]], summary: list[dict[str, str]]) -> None:
    wb = Workbook()

    ws_assigned = wb.active
    ws_assigned.title = "Assignments"
    assigned_headers = [
        "event_id",
        "date",
        "slot",
        "store",
        "brand",
        "category",
        "employee_id",
        "employee_name",
        "seat_no",
    ]
    ws_assigned.append(assigned_headers)
    for row in assignments:
        ws_assigned.append([row[header] for header in assigned_headers])

    ws_unfilled = wb.create_sheet("Unfilled")
    unfilled_headers = ["event_id", "date", "slot", "store", "brand", "category", "missing_seat", "reason"]
    ws_unfilled.append(unfilled_headers)
    for row in unfilled:
        ws_unfilled.append([row[header] for header in unfilled_headers])

    ws_summary = wb.create_sheet("Summary")
    summary_headers = ["employee_id", "name", "assigned_shifts", "weekly_max_shifts", "groups"]
    ws_summary.append(summary_headers)
    for row in summary:
        ws_summary.append([row[header] for header in summary_headers])

    wb.save(path)


def main() -> int:
    if len(sys.argv) != 3:
        print("Usage: python3 scheduler.py <input.xlsx> <output.xlsx>")
        return 1

    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        return 1

    assignments, unfilled, summary = run_scheduler(input_path, output_path)

    print(f"Created {output_path}")
    print(f"Assignments: {len(assignments)}")
    print(f"Unfilled seats: {len(unfilled)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
